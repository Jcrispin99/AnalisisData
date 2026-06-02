import io
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django import forms
from django.conf import settings
from django.contrib import admin, messages
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.urls import reverse
from django.utils.html import format_html
from django.utils.safestring import mark_safe

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from unfold.admin import ModelAdmin
from unfold.decorators import action

from .models import Atencion, AtencionNutricion, Paciente


EXCEL_COLUMNS = [
    ("id_rh", "Id RH"),
    ("nombre_completo", "Nombre Completo"),
    ("fec_ingreso", "Fec. Ingreso"),
    ("nro_documento", "Nº de documento"),
    ("convenio", "Convenio"),
    ("nombre_posicion", "Nombre posición"),
    ("nombre_area", "Nombre de Area"),
    ("nombre_unidad_org", "Nombre unidad org."),
    ("fec_nacimiento", "Fec. Nacimiento"),
    ("sexo", "Sexo"),
    ("instruccion", "Instrucción"),
    ("departamento", "Departamento"),
    ("provincia", "Provincia"),
    ("distrito", "Distrito"),
    ("direccion", "Dirección"),
]

DATE_FIELDS = {"fec_ingreso", "fec_nacimiento"}

# Mapea variantes en el Excel (acentos/mayúsculas se ignoran, ver _normalize_name)
# al valor de choice del modelo. Si la celda viene vacía se deja "" (sin clasificar).
CHOICE_NORMALIZERS = {
    "sexo": {
        "M": "M", "MASCULINO": "M", "VARON": "M", "HOMBRE": "M",
        "F": "F", "FEMENINO": "F", "MUJER": "F",
    },
    "instruccion": {
        "SIN INSTRUCCION": "SIN_INSTRUCCION",
        "SIN_INSTRUCCION": "SIN_INSTRUCCION",
        "NINGUNA": "SIN_INSTRUCCION",
        "PRIMARIA": "PRIMARIA",
        "SECUNDARIA": "SECUNDARIA",
        "TECNICA": "TECNICA",
        "TECNICO": "TECNICA",
        "SUPERIOR": "SUPERIOR",
        "UNIVERSITARIA": "SUPERIOR",
        "UNIVERSITARIO": "SUPERIOR",
        "POSGRADO": "POSGRADO",
        "POSTGRADO": "POSGRADO",
        "MAESTRIA": "POSGRADO",
        "DOCTORADO": "POSGRADO",
    },
}


ATENCION_COLUMNS = [
    ("nombre_completo", "Nombre Completo"),
    ("dni", "DNI Paciente"),
    ("fecha", "Fecha"),
    ("peso", "Peso (kg)"),
    ("talla", "Talla (cm)"),
    ("imc", "IMC"),
    ("sistolica", "Sistólica"),
    ("diastolica", "Diastólica"),
    ("abdominal", "Abdominal (cm)"),
    ("icc", "ICC"),
    ("colesterol_total", "Colesterol Total"),
    ("hdl", "HDL Colesterol"),
    ("ldl", "LDL Colesterol"),
    ("vldl", "VLDL Colesterol"),
    ("trigliceridos", "Triglicéridos"),
    ("hemoglobina", "Hemoglobina"),
    ("hematocrito", "Hematocrito"),
    ("glucosa", "Glucosa"),
    ("hb_a1c", "HB A1c"),
]

ATENCION_NUTRICION_COLUMNS = [
    ("nombre_completo", "Nombre Completo"),
    ("dni", "DNI Paciente"),
    ("fecha", "Fecha"),
    ("peso", "Peso (kg)"),
    ("talla", "Talla (cm)"),
    ("imc", "IMC"),
    ("grasa_corporal", "% Grasa corporal"),
    ("grasa_visceral", "% Grasa visceral"),
    ("masa_muscular", "% Masa muscular"),
]

PACIENTE_LOOKUP_FIELDS = {"dni", "nombre_completo"}


class ExcelUploadForm(forms.Form):
    archivo = forms.FileField(
        label="Archivo Excel (.xlsx)",
        help_text="Use la plantilla descargada para conservar los nombres de columna.",
    )


def _blankish(value):
    """True si la celda es None, vacía o contiene sólo whitespace.

    Cubre espacio normal, tabs, saltos de línea y espacio duro (`\\xa0`).
    Los Excel exportados desde sistemas clínicos a menudo rellenan nulos
    con un único espacio en vez de dejar la celda realmente vacía.
    """
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _parse_date(value, month_first=False):
    if _blankish(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if month_first:
        formats = ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%d/%m/%Y", "%d-%m-%Y")
    else:
        formats = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%m-%d-%Y")
    for fmt in formats:
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    raise ValueError(f"fecha no reconocida: {value!r}")


def _parse_decimal(value):
    if _blankish(value):
        return None
    try:
        return Decimal(str(value).strip().replace(",", "."))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"número no reconocido: {value!r}") from exc


def _build_template_response(columns, sheet_name, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    for col_idx, (_, label) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(label) + 4, 18)

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


_PUNCT_TO_SPACE = str.maketrans({c: " " for c in ",.;:-_/"})


def _normalize_name(s):
    """Normaliza un nombre para comparación.

    Aplica, en orden:
    - Quita acentos (NFKD).
    - Reemplaza puntuación común por espacios (la coma del "Apellidos, Nombres"
      es la causa típica de no-match contra Excel que viene sin coma).
    - Colapsa espacios consecutivos.
    - MAYÚSCULAS.

    "Sandoval Carbonell, Alfonso" → "SANDOVAL CARBONELL ALFONSO"
    "SANDOVAL CARBONELL ALFONSO"  → "SANDOVAL CARBONELL ALFONSO"
    "  José  García  "            → "JOSE GARCIA"
    """
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.translate(_PUNCT_TO_SPACE)
    return " ".join(s.split()).upper()


def _build_pacientes_index():
    """Devuelve dict {nombre_normalizado: [pacientes]} para lookup acento-insensible."""
    index = {}
    for p in Paciente.objects.only("id", "nro_documento", "nombre_completo"):
        index.setdefault(_normalize_name(p.nombre_completo), []).append(p)
    return index


def _buscar_paciente(dni, nombre, pacientes_by_norm):
    """Resuelve un Paciente por DNI (prioridad) o Nombre Completo.

    Reglas:
    - Si hay DNI: debe coincidir exactamente (no hay fallback silencioso al nombre).
    - Si no hay DNI pero sí nombre: comparación ignora acentos, mayúsculas y espacios extra.
      Falla si hay 0 o >1 coincidencias.

    Devuelve (paciente_o_None, mensaje_de_error_o_None).
    """
    if dni:
        try:
            return Paciente.objects.get(nro_documento=dni), None
        except Paciente.DoesNotExist:
            return None, f"paciente con DNI {dni!r} no existe."
        except Paciente.MultipleObjectsReturned:
            return None, f"DNI {dni!r} corresponde a más de un paciente."

    if not nombre:
        return None, "se requiere 'Nombre Completo' o 'DNI Paciente'."

    matches = pacientes_by_norm.get(_normalize_name(nombre), [])
    nombre_visible = " ".join(nombre.split())
    if not matches:
        return None, f"paciente con nombre {nombre_visible!r} no existe."
    if len(matches) > 1:
        return None, (
            f"nombre {nombre_visible!r} es ambiguo: {len(matches)} pacientes coinciden. "
            "Agregue el DNI para desambiguar."
        )
    return matches[0], None


def _diag_value(value):
    """Devuelve una descripción detallada de un valor de celda Excel para diagnóstico.

    Incluye tipo, repr, longitud (si es str) y los códigos Unicode de los
    primeros 10 caracteres. Permite detectar espacios duros (`\\xa0`), apóstrofos
    de texto, BOM, tabs, etc. en celdas que visualmente parecen vacías.
    """
    t = type(value).__name__
    info = f"type={t} repr={value!r}"
    if isinstance(value, str):
        codes = [ord(c) for c in value[:10]]
        info += f" len={len(value)} ord_first10={codes}"
    return info


class _ImportLog:
    """Log por-subida para diagnosticar fallos de carga Excel.

    Crea un archivo en `<BASE_DIR>/logs/import_<ts>_<modelo>.log` y guarda una
    copia del archivo subido en `<BASE_DIR>/logs/uploads/<ts>_<nombre>`. Permite
    pegar el log al técnico para que sepa exactamente qué celda rompió.
    """

    def __init__(self, archivo, modelo_label):
        base = Path(settings.BASE_DIR)
        logs_dir = base / "logs"
        uploads_dir = logs_dir / "uploads"
        logs_dir.mkdir(parents=True, exist_ok=True)
        uploads_dir.mkdir(parents=True, exist_ok=True)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        orig_name = getattr(archivo, "name", "upload.xlsx") or "upload.xlsx"
        safe_name = "".join(
            c if c.isalnum() or c in "._-" else "_" for c in orig_name
        )[:80]

        try:
            archivo.seek(0)
            raw_bytes = archivo.read()
            archivo.seek(0)
        except Exception:
            raw_bytes = b""

        self.upload_copy = uploads_dir / f"{ts}_{safe_name}"
        if raw_bytes:
            self.upload_copy.write_bytes(raw_bytes)

        self.path = logs_dir / f"import_{ts}_{modelo_label}.log"
        self._fh = open(self.path, "w", encoding="utf-8")
        self.error_count = 0

        self.write(f"=== Import {ts} ({modelo_label}) ===")
        self.write(f"archivo: {orig_name}")
        self.write(f"copia: {self.upload_copy}")
        self.write(f"tamano_bytes: {len(raw_bytes)}")

    def write(self, line):
        self._fh.write(line + "\n")
        self._fh.flush()

    def headers(self, headers, field_to_col):
        self.write("")
        self.write("HEADERS encontrados (columna Excel -> header -> campo modelo):")
        for col_letter, header, field in field_to_col:
            self.write(f"  {col_letter}: {header!r} -> {field}")
        self.write("")

    def cell_error(self, row_num, field_name, col_letter, value, exc):
        self.error_count += 1
        cell_ref = f"{col_letter}{row_num}" if col_letter else f"fila {row_num}"
        self.write(
            f"ERROR cell={cell_ref} campo={field_name} | {_diag_value(value)} | excepcion={exc}"
        )

    def row_error(self, row_num, msg):
        self.error_count += 1
        self.write(f"ERROR fila {row_num} | {msg}")

    def summary(self, creados, actualizados, errores):
        self.write("")
        self.write(
            f"=== Resumen: creados={creados} actualizados={actualizados} "
            f"errores={len(errores)} ==="
        )

    def close(self):
        try:
            self._fh.close()
        except Exception:
            pass


def _procesar_excel_atenciones(archivo, model, columns):
    """Upsert masivo de atenciones a partir de un xlsx con la plantilla dada.

    Vincula al paciente por Nombre Completo o DNI (ver `_buscar_paciente`).
    Clave natural del upsert: (paciente, fecha). Devuelve (creados, actualizados, errores).
    """
    log = _ImportLog(archivo, model._meta.model_name)
    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
        ws = wb.active

        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            log.write("FATAL: archivo vacío")
            raise ValueError("El archivo está vacío.")

        label_to_field = {label: field for field, label in columns}
        field_order = []
        col_letters = []
        header_field_pairs = []
        for col_idx, h in enumerate(headers, start=1):
            if h is None:
                continue
            if h not in label_to_field:
                log.write(f"FATAL: columna desconocida {h!r} en {get_column_letter(col_idx)}")
                raise ValueError(
                    f"Columna desconocida: {h!r}. Use la plantilla descargada."
                )
            f = label_to_field[h]
            field_order.append(f)
            col_letters.append(get_column_letter(col_idx))
            header_field_pairs.append((get_column_letter(col_idx), h, f))

        log.headers(headers, header_field_pairs)

        decimal_fields = {f for f, _ in columns} - PACIENTE_LOOKUP_FIELDS - {"fecha"}

        pacientes_by_norm = _build_pacientes_index()

        creados = actualizados = 0
        errores = []

        for row_num, row in enumerate(rows, start=2):
            if not row or all(_blankish(cell) for cell in row):
                continue

            data = {}
            row_ok = True
            for field_name, col_letter, value in zip(field_order, col_letters, row):
                if field_name == "fecha":
                    try:
                        data["fecha"] = _parse_date(value)
                    except ValueError as exc:
                        errores.append(f"Fila {row_num} (Fecha): {exc}")
                        log.cell_error(row_num, "fecha", col_letter, value, exc)
                        row_ok = False
                        break
                elif field_name in PACIENTE_LOOKUP_FIELDS:
                    data[field_name] = "" if value is None else str(value).strip()
                elif field_name in decimal_fields:
                    try:
                        data[field_name] = _parse_decimal(value)
                    except ValueError as exc:
                        errores.append(f"Fila {row_num} ({field_name}): {exc}")
                        log.cell_error(row_num, field_name, col_letter, value, exc)
                        row_ok = False
                        break

            if not row_ok:
                continue

            dni = data.pop("dni", "")
            nombre = data.pop("nombre_completo", "")
            paciente, error = _buscar_paciente(dni, nombre, pacientes_by_norm)
            if error:
                errores.append(f"Fila {row_num}: {error}")
                log.row_error(row_num, error)
                continue

            fecha = data.pop("fecha", None)
            if not fecha:
                errores.append(f"Fila {row_num}: 'Fecha' vacía, fila omitida.")
                log.row_error(row_num, "'Fecha' vacía, fila omitida.")
                continue

            _, created = model.objects.update_or_create(
                paciente=paciente, fecha=fecha, defaults=data
            )
            if created:
                creados += 1
            else:
                actualizados += 1

        log.summary(creados, actualizados, errores)
        return creados, actualizados, errores, log.path
    finally:
        log.close()


@admin.register(Paciente)
class PacienteAdmin(ModelAdmin):
    list_display = (
        "id_rh",
        "nombre_completo",
        "nro_documento",
        "convenio",
        "nombre_area",
        "fec_ingreso",
    )
    list_filter = ("convenio", "nombre_area", "departamento", "provincia", "distrito")
    search_fields = ("id_rh", "nombre_completo", "nro_documento")
    ordering = ("nombre_completo",)
    fieldsets = (
        ("Identificación", {
            "fields": (
                "id_rh",
                "nombre_completo",
                "nro_documento",
                "fec_nacimiento",
                "sexo",
                "instruccion",
            ),
        }),
        ("Información laboral", {
            "fields": (
                "fec_ingreso",
                "convenio",
                "nombre_posicion",
                "nombre_area",
                "nombre_unidad_org",
            ),
        }),
        ("Ubicación", {
            "fields": ("departamento", "provincia", "distrito", "direccion", "altitud_msnm"),
        }),
    )

    actions_list = ("descargar_plantilla", "subir_excel")

    @action(description="Descargar plantilla", url_path="descargar-plantilla")
    def descargar_plantilla(self, request):
        return _build_template_response(
            EXCEL_COLUMNS, "Pacientes", "plantilla_pacientes.xlsx"
        )

    @action(description="Subir Excel", url_path="subir-excel")
    def subir_excel(self, request):
        cancel_url = reverse("admin:pacientes_paciente_changelist")
        if request.method == "POST":
            form = ExcelUploadForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    creados, actualizados, errores, log_path = self._procesar_excel(
                        form.cleaned_data["archivo"]
                    )
                except Exception as exc:
                    messages.error(request, f"Error procesando el archivo: {exc}")
                    return HttpResponseRedirect(request.path)

                for err in errores[:10]:
                    messages.warning(request, err)
                if len(errores) > 10:
                    messages.warning(request, f"... y {len(errores) - 10} errores más.")

                messages.info(request, f"Log detallado: {log_path}")
                messages.success(
                    request,
                    f"Carga completa. Creados: {creados}, actualizados: {actualizados}, "
                    f"con errores: {len(errores)}.",
                )
                return HttpResponseRedirect(cancel_url)
        else:
            form = ExcelUploadForm()

        context = {
            **self.admin_site.each_context(request),
            "form": form,
            "title": "Subir Pacientes desde Excel",
            "opts": self.model._meta,
            "columnas": [label for _, label in EXCEL_COLUMNS],
            "cancel_url": cancel_url,
            "key_fields_help": (
                "El campo <strong>Id RH</strong> es obligatorio y se usa para "
                "identificar al registro (si ya existe, se actualiza)."
            ),
        }
        return render(request, "admin/pacientes/subir_excel.html", context)

    def _procesar_excel(self, archivo):
        log = _ImportLog(archivo, "paciente")
        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb.active

            rows = ws.iter_rows(values_only=True)
            headers = next(rows, None)
            if not headers:
                log.write("FATAL: archivo vacío")
                raise ValueError("El archivo está vacío.")

            label_to_field = {label: field for field, label in EXCEL_COLUMNS}
            field_order = []
            col_letters = []
            header_field_pairs = []
            for col_idx, h in enumerate(headers, start=1):
                if h is None:
                    continue
                if h not in label_to_field:
                    log.write(f"FATAL: columna desconocida {h!r} en {get_column_letter(col_idx)}")
                    raise ValueError(
                        f"Columna desconocida: {h!r}. Use la plantilla descargada."
                    )
                f = label_to_field[h]
                field_order.append(f)
                col_letters.append(get_column_letter(col_idx))
                header_field_pairs.append((get_column_letter(col_idx), h, f))

            log.headers(headers, header_field_pairs)

            creados = actualizados = 0
            errores = []

            for row_num, row in enumerate(rows, start=2):
                if not row or all(_blankish(cell) for cell in row):
                    continue

                data = {}
                row_ok = True
                for field_name, col_letter, value in zip(field_order, col_letters, row):
                    if field_name in DATE_FIELDS:
                        try:
                            data[field_name] = _parse_date(
                                value, month_first=(field_name == "fec_nacimiento")
                            )
                        except ValueError as exc:
                            errores.append(f"Fila {row_num} ({field_name}): {exc}")
                            log.cell_error(row_num, field_name, col_letter, value, exc)
                            row_ok = False
                            break
                    elif field_name in CHOICE_NORMALIZERS:
                        if _blankish(value):
                            data[field_name] = ""
                        else:
                            normalized = _normalize_name(value)
                            key = CHOICE_NORMALIZERS[field_name].get(normalized)
                            if key is None:
                                msg = (
                                    f"valor {str(value)!r} no reconocido, "
                                    f"esperado uno de {sorted(set(CHOICE_NORMALIZERS[field_name].values()))}."
                                )
                                errores.append(f"Fila {row_num} ({field_name}): {msg}")
                                log.cell_error(row_num, field_name, col_letter, value, msg)
                                row_ok = False
                                break
                            data[field_name] = key
                    else:
                        data[field_name] = "" if value is None else str(value).strip()

                if not row_ok:
                    continue

                id_rh = data.get("id_rh", "").strip()
                if not id_rh:
                    errores.append(f"Fila {row_num}: 'Id RH' vacío, fila omitida.")
                    log.row_error(row_num, "'Id RH' vacío, fila omitida.")
                    continue

                data.pop("id_rh")
                _, created = Paciente.objects.update_or_create(
                    id_rh=id_rh, defaults=data
                )
                if created:
                    creados += 1
                else:
                    actualizados += 1

            log.summary(creados, actualizados, errores)
            return creados, actualizados, errores, log.path
        finally:
            log.close()


class _AtencionImportExportMixin:
    """Acciones compartidas para descargar plantilla y subir Excel.

    Subclases deben definir: `import_columns`, `import_sheet_name`,
    `import_filename`, `import_title`.
    """

    import_columns = ()
    import_sheet_name = ""
    import_filename = ""
    import_title = ""

    actions_list = ("descargar_plantilla", "subir_excel")

    def _changelist_url(self):
        opts = self.model._meta
        return reverse(f"admin:{opts.app_label}_{opts.model_name}_changelist")

    @action(description="Descargar plantilla", url_path="descargar-plantilla")
    def descargar_plantilla(self, request):
        return _build_template_response(
            self.import_columns, self.import_sheet_name, self.import_filename
        )

    @action(description="Subir Excel", url_path="subir-excel")
    def subir_excel(self, request):
        cancel_url = self._changelist_url()
        if request.method == "POST":
            form = ExcelUploadForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    creados, actualizados, errores, log_path = _procesar_excel_atenciones(
                        form.cleaned_data["archivo"], self.model, self.import_columns
                    )
                except Exception as exc:
                    messages.error(request, f"Error procesando el archivo: {exc}")
                    return HttpResponseRedirect(request.path)

                for err in errores[:10]:
                    messages.warning(request, err)
                if len(errores) > 10:
                    messages.warning(request, f"... y {len(errores) - 10} errores más.")

                messages.info(request, f"Log detallado: {log_path}")
                messages.success(
                    request,
                    f"Carga completa. Creados: {creados}, actualizados: {actualizados}, "
                    f"con errores: {len(errores)}.",
                )
                return HttpResponseRedirect(cancel_url)
        else:
            form = ExcelUploadForm()

        context = {
            **self.admin_site.each_context(request),
            "form": form,
            "title": self.import_title,
            "opts": self.model._meta,
            "columnas": [label for _, label in self.import_columns],
            "cancel_url": cancel_url,
            "key_fields_help": (
                "El paciente se identifica por <strong>Nombre Completo</strong> "
                "(la comparación ignora acentos, mayúsculas y espacios extra: "
                "<code>José Pérez</code> = <code>JOSE PEREZ</code>) o por "
                "<strong>DNI Paciente</strong>. Si llenas DNI, se usa como prioridad. "
                "Si el nombre coincide con más de un paciente, agrega el DNI para "
                "desambiguar. <strong>Fecha</strong> es obligatoria y forma la clave "
                "única del registro junto con el paciente (si ya existe una atención "
                "de ese paciente en esa fecha, se actualiza)."
            ),
        }
        return render(request, "admin/pacientes/subir_excel.html", context)


# Colores para badges de clasificación (compatibles con Tailwind/Unfold)
_BADGE_COLORS = {
    # neutros
    "Normal": "bg-green-100 text-green-800",
    "Deseable": "bg-green-100 text-green-800",
    # leves
    "Dislipidemia leve": "bg-yellow-100 text-yellow-800",
    "Eritrocitosis leve": "bg-yellow-100 text-yellow-800",
    "Prediabetes": "bg-yellow-100 text-yellow-800",
    "Presión elevada": "bg-yellow-100 text-yellow-800",
    "Grado 1 (riesgo aumentado)": "bg-yellow-100 text-yellow-800",
    "Hipertensión grado 1": "bg-yellow-100 text-yellow-800",
    # moderadas
    "Dislipidemia moderada": "bg-orange-100 text-orange-800",
    "Eritrocitosis moderada": "bg-orange-100 text-orange-800",
    "Grado 2 (riesgo significativo)": "bg-orange-100 text-orange-800",
    "Hipertensión grado 2": "bg-orange-100 text-orange-800",
    # severas
    "Dislipidemia severa": "bg-red-100 text-red-800",
    "Eritrocitosis severa": "bg-red-100 text-red-800",
    "Diabetes": "bg-red-100 text-red-800",
    "Crisis hipertensiva": "bg-red-100 text-red-800",
    # diagnóstico combinado de dislipidemia
    "Perfil lipídico normal": "bg-green-100 text-green-800",
    "HDL bajo aislado": "bg-yellow-100 text-yellow-800",
    "Hipercolesterolemia leve": "bg-yellow-100 text-yellow-800",
    "Hipertrigliceridemia leve": "bg-yellow-100 text-yellow-800",
    "Dislipidemia mixta leve": "bg-yellow-100 text-yellow-800",
    "Hipercolesterolemia moderada": "bg-orange-100 text-orange-800",
    "Hipertrigliceridemia moderada": "bg-orange-100 text-orange-800",
    "Dislipidemia mixta moderada": "bg-orange-100 text-orange-800",
    "Hipercolesterolemia severa": "bg-red-100 text-red-800",
    "Hipertrigliceridemia severa": "bg-red-100 text-red-800",
    "Dislipidemia mixta severa": "bg-red-100 text-red-800",
    "Dislipidemia aterogénica": "bg-red-100 text-red-800",
    # colesterol no-HDL
    "Limítrofe alto": "bg-yellow-100 text-yellow-800",
    "Riesgo aumentado": "bg-orange-100 text-orange-800",
    "Riesgo muy alto": "bg-red-100 text-red-800",
    # composición corporal — IMC
    "Bajo peso": "bg-yellow-100 text-yellow-800",
    "Sobrepeso": "bg-yellow-100 text-yellow-800",
    "Obesidad grado I": "bg-orange-100 text-orange-800",
    "Obesidad grado II": "bg-red-100 text-red-800",
    "Obesidad grado III": "bg-red-100 text-red-800",
    # composición corporal — grasa / músculo
    "Saludable": "bg-green-100 text-green-800",
    "Aceptable": "bg-yellow-100 text-yellow-800",
    "Obesidad": "bg-red-100 text-red-800",
    "Alta": "bg-red-100 text-red-800",
    "Bajo": "bg-yellow-100 text-yellow-800",
    "Alto": "bg-green-100 text-green-800",
    "Muy alto": "bg-blue-100 text-blue-800",
}


def _badge(label):
    if not label:
        return mark_safe('<span class="text-font-subtle-light dark:text-font-subtle-dark">—</span>')
    css = _BADGE_COLORS.get(label, "bg-base-100 text-base-700 dark:bg-base-800 dark:text-base-200")
    return format_html(
        '<span class="inline-block px-2 py-0.5 rounded-full text-xs font-medium {}">{}</span>',
        css,
        label,
    )


@admin.register(Atencion)
class AtencionAdmin(_AtencionImportExportMixin, ModelAdmin):
    autocomplete_fields = ("paciente",)
    date_hierarchy = "fecha"
    list_display = (
        "fecha",
        "paciente_dni",
        "paciente_nombre",
        "imc",
        "presion_badge",
        "obesidad_badge",
        "glucosa_badge",
        "hb_a1c_badge",
        "dislipidemia_badge",
    )
    list_filter = ("fecha", "paciente__convenio", "paciente__nombre_area")
    search_fields = (
        "paciente__nro_documento",
        "paciente__nombre_completo",
        "paciente__id_rh",
    )
    ordering = ("-fecha",)
    readonly_fields = ("clasificaciones_resumen",)
    fieldsets = (
        ("Paciente", {
            "fields": ("paciente", "fecha"),
        }),
        ("Antropometría", {
            "fields": ("peso", "talla", "imc", "abdominal", "icc"),
        }),
        ("Presión arterial", {
            "fields": ("sistolica", "diastolica"),
        }),
        ("Perfil lipídico", {
            "fields": ("colesterol_total", "hdl", "ldl", "vldl", "trigliceridos"),
        }),
        ("Hematología y glucosa", {
            "fields": ("hemoglobina", "hematocrito", "glucosa", "hb_a1c"),
        }),
        ("Clasificación clínica (calculada)", {
            "fields": ("clasificaciones_resumen",),
        }),
    )

    import_columns = ATENCION_COLUMNS
    import_sheet_name = "Atenciones"
    import_filename = "plantilla_atenciones.xlsx"
    import_title = "Subir Atenciones desde Excel"

    def get_queryset(self, request):
        # Reduce queries al renderizar badges que necesitan paciente.sexo y paciente.es_altura.
        return super().get_queryset(request).select_related("paciente")

    @admin.display(description="DNI", ordering="paciente__nro_documento")
    def paciente_dni(self, obj):
        return obj.paciente.nro_documento

    @admin.display(description="Paciente", ordering="paciente__nombre_completo")
    def paciente_nombre(self, obj):
        return obj.paciente.nombre_completo

    @admin.display(description="Presión")
    def presion_badge(self, obj):
        return _badge(obj.clasif_presion)

    @admin.display(description="Abdominal")
    def obesidad_badge(self, obj):
        return _badge(obj.clasif_obesidad_abdominal)

    @admin.display(description="Glicemia")
    def glucosa_badge(self, obj):
        return _badge(obj.clasif_glicemia_ayunas)

    @admin.display(description="HbA1c")
    def hb_a1c_badge(self, obj):
        return _badge(obj.clasif_hb_a1c)

    @admin.display(description="Col. Total")
    def colesterol_badge(self, obj):
        return _badge(obj.clasif_colesterol_total)

    @admin.display(description="Dislipidemia")
    def dislipidemia_badge(self, obj):
        return _badge(obj.clasif_dislipidemia)

    @admin.display(description="Resumen de clasificaciones")
    def clasificaciones_resumen(self, obj):
        no_hdl = obj.colesterol_no_hdl
        no_hdl_label = (
            format_html("{} mg/dL — {}", no_hdl, _badge(obj.clasif_colesterol_no_hdl))
            if no_hdl is not None
            else _badge(None)
        )
        dislip_cell = _badge(obj.clasif_dislipidemia)
        if obj.sospecha_hipercolesterolemia_familiar:
            dislip_cell = format_html(
                "{} <span class='text-xs text-red-700 dark:text-red-300 ml-1'>"
                "⚠ sospecha hipercolesterolemia familiar (LDL ≥190)</span>",
                dislip_cell,
            )
        filas = [
            ("Diagnóstico de dislipidemia", dislip_cell),
            ("Colesterol Total", _badge(obj.clasif_colesterol_total)),
            ("HDL", _badge(obj.clasif_hdl)),
            ("LDL", _badge(obj.clasif_ldl)),
            ("Triglicéridos", _badge(obj.clasif_trigliceridos)),
            ("Colesterol no-HDL", no_hdl_label),
            ("Eritrocitosis (Hb)", _badge(obj.clasif_eritrocitosis)),
            ("Obesidad abdominal", _badge(obj.clasif_obesidad_abdominal)),
            ("Glicemia en ayunas", _badge(obj.clasif_glicemia_ayunas)),
            ("Hemoglobina glicosilada", _badge(obj.clasif_hb_a1c)),
            ("Presión arterial", _badge(obj.clasif_presion)),
        ]
        rows = "".join(
            format_html(
                '<tr><td class="pr-4 py-1 font-medium">{}</td><td class="py-1">{}</td></tr>',
                nombre,
                celda,
            )
            for nombre, celda in filas
        )
        altura_nota = ""
        if obj.paciente_id:
            altitud = obj.paciente.altitud_efectiva
            if altitud is None:
                altura_nota = "<p class='text-xs mt-2 text-font-subtle-light dark:text-font-subtle-dark'>Altitud no determinada — eritrocitosis usa tabla &lt;2500 msnm.</p>"
            else:
                altura_nota = format_html(
                    "<p class='text-xs mt-2 text-font-subtle-light dark:text-font-subtle-dark'>Altitud estimada: {} msnm ({}).</p>",
                    altitud,
                    "≥2500" if obj.paciente.es_altura else "<2500",
                )
        return format_html("<table>{}</table>{}", mark_safe(rows), mark_safe(altura_nota))


@admin.register(AtencionNutricion)
class AtencionNutricionAdmin(_AtencionImportExportMixin, ModelAdmin):
    autocomplete_fields = ("paciente",)
    date_hierarchy = "fecha"
    list_display = (
        "fecha",
        "paciente_dni",
        "paciente_nombre",
        "imc",
        "imc_badge",
        "grasa_corporal_badge",
        "grasa_visceral_badge",
        "masa_muscular_badge",
    )
    list_filter = ("fecha", "paciente__convenio", "paciente__nombre_area")
    search_fields = (
        "paciente__nro_documento",
        "paciente__nombre_completo",
        "paciente__id_rh",
    )
    ordering = ("-fecha",)
    readonly_fields = ("clasificaciones_resumen",)
    fieldsets = (
        ("Paciente", {
            "fields": ("paciente", "fecha"),
        }),
        ("Antropometría", {
            "fields": ("peso", "talla", "imc"),
        }),
        ("Composición corporal", {
            "fields": ("grasa_corporal", "grasa_visceral", "masa_muscular"),
        }),
        ("Clasificación (calculada)", {
            "fields": ("clasificaciones_resumen",),
        }),
    )

    def get_queryset(self, request):
        return super().get_queryset(request).select_related("paciente")

    @admin.display(description="IMC")
    def imc_badge(self, obj):
        return _badge(obj.clasif_imc)

    @admin.display(description="% Grasa corp.")
    def grasa_corporal_badge(self, obj):
        return _badge(obj.clasif_grasa_corporal)

    @admin.display(description="Grasa visceral")
    def grasa_visceral_badge(self, obj):
        return _badge(obj.clasif_grasa_visceral)

    @admin.display(description="% Masa muscular")
    def masa_muscular_badge(self, obj):
        return _badge(obj.clasif_masa_muscular)

    @admin.display(description="Clasificación")
    def clasificaciones_resumen(self, obj):
        if not obj.pk:
            return "—"
        items = [
            ("IMC", obj.clasif_imc),
            ("% Grasa corporal", obj.clasif_grasa_corporal),
            ("Grasa visceral", obj.clasif_grasa_visceral),
            ("% Masa muscular", obj.clasif_masa_muscular),
        ]
        rows = ""
        for label, valor in items:
            rows += format_html(
                "<tr><td style='padding:2px 8px;'><strong>{}</strong></td>"
                "<td style='padding:2px 8px;'>{}</td></tr>",
                label,
                _badge(valor),
            )
        return format_html("<table>{}</table>", mark_safe(rows))

    import_columns = ATENCION_NUTRICION_COLUMNS
    import_sheet_name = "Nutrición"
    import_filename = "plantilla_atenciones_nutricion.xlsx"
    import_title = "Subir Atenciones de Nutrición desde Excel"

    @admin.display(description="DNI", ordering="paciente__nro_documento")
    def paciente_dni(self, obj):
        return obj.paciente.nro_documento

    @admin.display(description="Paciente", ordering="paciente__nombre_completo")
    def paciente_nombre(self, obj):
        return obj.paciente.nombre_completo

